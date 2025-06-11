import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from openpyxl.styles import Font


# --- Load Excel ---
df = pd.read_excel("usaha.xlsx")
idsbr_list = df['idsbr mirip'].dropna().astype(str).tolist()

# --- Setup Selenium dengan Chrome profile (agar sudah login SSO dan VPN aktif) ---
options = webdriver.ChromeOptions()
chrome_profile_path = r"C:\\Shared\\Coding\\script\\bot_profil"
options.add_argument(f"--user-data-dir={chrome_profile_path}")
driver = webdriver.Chrome(options=options)

# --- Buka halaman utama ---
driver.get("https://matchapro.web.bps.go.id/direktori-usaha")
time.sleep(7)  # Tunggu halaman dan JS selesai load

# Load workbook dan worksheet sekali saja
wb = load_workbook("usaha.xlsx")
ws = wb.active

# Buat set berisi nomor baris yang sudah bold
bold_rows = set()
for row in ws.iter_rows(min_row=2):  # skip header
    if any(cell.font and cell.font.bold for cell in row):
        bold_rows.add(row[0].row)  # baris ke-i (misalnya 2, 3, dst)

for i, row in reversed(list(df.iterrows())):
    idsbr = str(row["idsbr mirip"])
    excel_row = i + 2  # baris Excel dimulai dari 2 (header di baris 1)
    if excel_row in bold_rows:
        print(f"⏩ Baris ke-{excel_row} sudah bold, dilewati.")
        continue

    try:
        print(f"Memproses IDSBR: {idsbr}")

        # Cari input dengan name=idsbr
        input_box = driver.find_element(By.NAME, "idsbr")
        input_box.clear()
        input_box.send_keys(idsbr)

        # Klik tombol filter
        filter_button = driver.find_element(By.ID, "filter-data")
        filter_button.click()

        time.sleep(7)  # Tunggu hasil filter keluar

        # Ambil semua tombol edit yang muncul
        edit_buttons = driver.find_elements(By.CSS_SELECTOR, 'a.btn-edit-perusahaan[aria-label="Edit"]')

        if len(edit_buttons) == 1:
            print(f"✅ Ditemukan 1 hasil, klik tombol edit untuk {idsbr}")
            # Klik tombol edit
            driver.execute_script("arguments[0].click();", edit_buttons[0])
            time.sleep(5)  # Tunggu Modal Entri
            print("🔄 Menunggu modal SweetAlert...")

            # Tunggu modal SweetAlert muncul dan klik tombol "Ya, edit!"
            try:
                # Tunggu tombol benar-benar visible dan aktif
                confirm_button = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, 'button.swal2-confirm.swal2-styled'))
                )

                # Klik pakai JavaScript (lebih paksa)
                driver.execute_script("arguments[0].click();", confirm_button)
                print("✅ Klik tombol 'Ya, edit!' berhasil (via JS).")
                print("✅ Klik tombol 'Ya, edit!' di modal berhasil.")
                
                # Deteksi halaman 'Profiling Info 🛠' (Not Authorized)
                try:
                    WebDriverWait(driver, 5).until(             
                        EC.presence_of_element_located((By.XPATH, "//h2[contains(text(), 'Profiling Info')]"))
                    )
                    print("🚫 Usaha sedang diedit orang lain / sudah terisi — lewati dan tandai.")

                    # Tandai di Excel
                    wb = load_workbook("usaha.xlsx")
                    ws = wb.active
                    ws[f"X{excel_row}"] = "Aji"
                    for cell in ws[excel_row]:
                        cell.font = Font(bold=True)
                    bold_rows.add(excel_row)    
                    wb.save("usaha.xlsx")

                    # Kembali ke halaman awal
                    driver.get("https://matchapro.web.bps.go.id/direktori-usaha")
                    time.sleep(7)
                    continue  # Lanjut ke baris selanjutnya
                except:
                    pass  # Jika tidak muncul, lanjutkan proses biasa

                #Entri Form
                # Tunggu form selesai load
                time.sleep(5)
                try:
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "submit-final"))
                    )
                except:
                    try:
                        WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.ID, "cancel-submit-final"))
                        )
                        print("🚫 Muncul tombol cancel-submit-final → lewati entri")
                        # Tandai sebagai diproses
                        wb = load_workbook("usaha.xlsx")
                        ws = wb.active
                        ws[f"X{excel_row}"] = "Aji"
                        for cell in ws[excel_row]:
                            cell.font = Font(bold=True)
                        wb.save("usaha.xlsx")
                        continue  # Langsung ke baris berikutnya
                    except:
                        print("❌ Tidak ditemukan submit-final atau cancel-submit-final")
                        continue

                time.sleep(5)
                # 1. Isi Alamat
                alamat_input = driver.find_element(By.ID, "alamat_usaha")
                alamat_input.clear()
                alamat_input.send_keys(row["alamat"])

                # 2. Isi Sumber Profiling
                try:                
                    sumber_input = driver.find_element(By.ID, "sumber_profiling")
                    sumber_input.clear()
                    sumber_input.send_keys(row["Sumber"])
                except:             
                    print("❌ Tidak ditemukan input sumber_profiling")

                # 3. Isi Catatan Profiling
                try:
                    catatan_input = driver.find_element(By.ID, "catatan_profiling")
                    catatan_input.clear()
                    catatan_input.send_keys(row["Desk Sumber"])
                except:
                    print("❌ Tidak ditemukan input catatan_profiling")

                # 4. Isi Latitude
                try:
                    lat_input = driver.find_element(By.ID, "latitude")
                    lat_input.clear()
                    lat_input.send_keys(str(row["latitude2"]))
                    time.sleep(1)
                except:
                    print("❌ Tidak ditemukan input latitude")

                # 5. Isi Longitude
                try:
                    lon_input = driver.find_element(By.ID, "longitude")
                    lon_input.clear()
                    lon_input.send_keys(str(row["longitude2"]))
                    time.sleep(1)
                except:
                    print("❌ Tidak ditemukan input longitude")

                # 6. Klik tombol "Cek Peta"
                try:
                    cek_peta_btn = driver.find_element(By.ID, "cek-peta")
                    driver.execute_script("arguments[0].click();", cek_peta_btn)
                    print("🗺️ Tombol 'Cek Peta' diklik.")
                    time.sleep(5)  # Delay setelah cek peta
                except:
                    print("❌ Tombol 'Cek Peta' tidak ditemukan")

                # 7. Email check/uncheck
                try:
                    email_input = driver.find_element(By.ID, "email")
                    if email_input.get_attribute("value").strip() == "":
                        # Kosong → uncheck
                        email_checkbox = driver.find_element(By.ID, "check-email")
                        if email_checkbox.is_selected():
                            driver.execute_script("arguments[0].click();", email_checkbox)
                            print("📭 Email kosong, checkbox di-uncheck.")
                except:
                    print("⚠️ Tidak bisa akses input email atau checkbox.")

                # 8. Klik tombol "Submit Final"
                try:
                    submit_btn = driver.find_element(By.ID, "submit-final")
                    driver.execute_script("arguments[0].click();", submit_btn)
                    print("✅ Tombol 'Submit Final' diklik.")
                    time.sleep(3)
                except:
                    print("❌ Gagal klik tombol 'Submit Final'")

                # 9. Jika muncul modal konfirmasi konsistensi, klik
                try:
                    confirm_modal_btn = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.ID, "confirm-consistency"))
                    )
                    driver.execute_script("arguments[0].click();", confirm_modal_btn)
                    print("✅ Modal konsistensi dikonfirmasi.")
                    time.sleep(2)
                except:
                    print("⚠️ Tidak ada modal confirm-consistency.")

                # 10. Jika muncul tombol "ignore-consistency", klik juga
                try:
                    ignore_modal_btn = WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((By.ID, "ignore-consistency"))
                    )
                    driver.execute_script("arguments[0].click();", ignore_modal_btn)
                    print("⚠️ Klik tombol 'ignore-consistency'")
                    time.sleep(2)
                except:
                    print("✅ Tidak perlu klik ignore-consistency.")

                # 11. Terakhir, klik tombol "Ya, Submit!" di dalam modal SweetAlert
                try:
                    final_submit_btn = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.swal2-confirm.btn.btn-primary"))
                    )
                    driver.execute_script("arguments[0].click();", final_submit_btn)
                    print("🎯 Klik tombol akhir 'Ya, Submit!' berhasil.")
                    time.sleep(3)
                except:
                    print("❌ Gagal menemukan atau klik tombol 'Ya, Submit!'")
                    
                # Tandai baris berhasil
                try:
                    # Buka workbook dan sheet
                    wb = load_workbook("usaha.xlsx")
                    ws = wb.active

                    # Baris Excel dimulai dari 2 karena baris 1 adalah header
                    excel_row = i + 2
                    ws[f"X{excel_row}"] = "Aji"  # Misal kolom Profiler ada di kolom G

                    # Bold seluruh baris
                    for cell in ws[excel_row]:
                        cell.font = Font(bold=True)

                    wb.save("usaha.xlsx")
                    print(f"📝 Baris ke-{i+1} ditandai sebagai selesai oleh Aji.")
                except Exception as e:
                    print(f"❌ Gagal update Excel: {e}")

            except:
                print("❌ Gagal menemukan atau klik tombol 'Ya, edit!'")
                time.sleep(10)
            continue  # Lanjut ke baris selanjutnya
        else:
            print(f"⚠️ Hasil IDSBR {idsbr} tidak unik atau tidak ditemukan.")

        # Reload ulang halaman utama untuk IDSBR berikutnya
        driver.get("https://matchapro.web.bps.go.id/direktori-usaha")
        time.sleep(7)

    except Exception as e:
        print(f"❌ Error saat memproses {idsbr}: {e}")
        driver.get("https://matchapro.web.bps.go.id/direktori-usaha")
        time.sleep(7)

driver.quit()
