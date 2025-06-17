import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from openpyxl.styles import Font
import time

# Load data Excel
df = pd.read_excel("usaha.xlsx")
idsbr_list = df['idsbr mirip'].dropna().astype(str).tolist()

# Setup Chrome dengan profile login
options = webdriver.ChromeOptions()
options.add_argument("--user-data-dir=C:\\Shared\\Coding\\script\\bot_profil")
driver = webdriver.Chrome(options=options)

# Load Excel
wb = load_workbook("usaha.xlsx")
ws = wb.active

# Fungsi bantu isi input
def isi_input(driver, by, locator, value):
    try:
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((by, locator)))
        elem = driver.find_element(by, locator)
        elem.clear()
        if value:
            elem.send_keys(str(value))
        print(f"✅ Input {locator} diisi.")
    except Exception as e:
        print(f"❌ Tidak bisa mengisi input dengan ID '{locator}' → {e}")

def safe_click(driver, by, locator):
    try:
        elem = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((by, locator)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", elem)
        time.sleep(0.5)  # beri waktu scroll
        elem.click()
        print(f"✅ Klik tombol {locator}")
        return True
    except Exception as e:
        print(f"❌ Gagal klik tombol {locator} → {e}")
        return False

def tunggu_loading_data_hilang(driver, max_wait1=15, max_wait2=10):
    print("⏳ Menunggu 'Loading Data' hilang (tahap 1)...")
    try:
        WebDriverWait(driver, max_wait1).until_not(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Loading Data')]"))
        )
        print("✅ 'Loading Data' hilang.")
        return True
    except:
        print(f"⚠️ Masih loading setelah {max_wait1} detik. Coba tunggu {max_wait2} detik lagi...")

    try:
        WebDriverWait(driver, max_wait2).until_not(
            EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Loading Data')]"))
        )
        print("✅ 'Loading Data' akhirnya hilang.")
        return True
    except:
        print(f"❌ 'Loading Data' masih ada setelah total {max_wait1 + max_wait2} detik.")
        return False
    
# Baris yang sudah dibold
bold_rows = set(row[0].row for row in ws.iter_rows(min_row=2) if any(cell.font and cell.font.bold for cell in row))

i = 0
max_retry = 3

while i < len(df):
    row = df.iloc[i]
    idsbr = str(int(row['idsbr mirip']))
    excel_row = i + 2

    if excel_row in bold_rows:
        print(f"⏩ Baris ke-{excel_row} sudah bold, dilewati.")
        i += 1
        continue

    retry_count = 0
    success = False

    while retry_count < max_retry and not success:
        print(f"\n🔁 Memproses baris {excel_row}, percobaan ke-{retry_count + 1}...")
        try:
            driver.get("https://matchapro.web.bps.go.id/direktori-usaha")
            WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.NAME, "idsbr")))
        except:
            print("❌ Gagal memuat halaman direktori, lanjut ID berikutnya.")
            continue

        try:
            print(f"Memproses IDSBR: {idsbr}")

            input_box = driver.find_element(By.NAME, "idsbr")
            input_box.clear()
            input_box.send_keys(idsbr)

            WebDriverWait(driver, 10).until_not(EC.presence_of_element_located((By.CLASS_NAME, "blockUI")))
            driver.find_element(By.ID, "filter-data").click()

            WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'a.btn-edit-perusahaan[aria-label="Edit"]')))
            edit_buttons = driver.find_elements(By.CSS_SELECTOR, 'a.btn-edit-perusahaan[aria-label="Edit"]')
            if len(edit_buttons) != 1:
                print(f"⚠️ IDSBR {idsbr} tidak unik atau tidak ditemukan.")
                continue

            driver.execute_script("arguments[0].click();", edit_buttons[0])
            print("🔄 Menunggu modal SweetAlert...")

            try:
                confirm_button = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, 'button.swal2-confirm.swal2-styled'))
                )
                driver.execute_script("arguments[0].click();", confirm_button)
                print("✅ Klik tombol 'Ya, edit!'")
            except:
                print("❌ Modal konfirmasi tidak muncul, lanjutkan")
                continue

            try:
                WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//h2[contains(text(), 'Profiling Info')]")))
                print("🚫 Sudah diedit oleh orang lain.")
                ws[f"X{excel_row}"] = "Aji"
                for cell in ws[excel_row]:
                    cell.font = Font(bold=True)
                wb.save("usaha.xlsx")
                continue
            except:
                pass

            try:
                submit_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "submit-final"))
                )
                is_visible = driver.execute_script(
                    """
                    var element = arguments[0];
                    return window.getComputedStyle(element).display !== 'none' && element.offsetParent !== null;
                    """, submit_element)
                if not is_visible:
                    print("🚫 Tombol 'submit-final' disembunyikan (display: none) → lewati entri.")
                    ws[f"X{excel_row}"] = "Aji"
                    for cell in ws[excel_row]:
                        cell.font = Font(bold=True)
                    wb.save("usaha.xlsx")
                    continue
            except:
                # Cek apakah cancel-submit-final muncul
                try:
                    WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "cancel-submit-final")))
                    print("🚫 Form tidak bisa diisi (cancel-submit-final).")
                    ws[f"X{excel_row}"] = "Aji"
                    for cell in ws[excel_row]:
                        cell.font = Font(bold=True)
                    wb.save("usaha.xlsx")
                    continue
                except:
                    print("❌ Tombol submit/cancel tidak terdeteksi.")
                    ws[f"X{excel_row}"] = "Aji - error form"
                    for cell in ws[excel_row]:
                        cell.font = Font(bold=True)
                    wb.save("usaha.xlsx")
                    continue

            print("📋 Daftar elemen yang tersedia:")
            # Tunggu hingga "Loading Data" hilang sebelum mulai input form
            try:
                print("⏳ Menunggu 'Loading Data' hilang...")
                WebDriverWait(driver, 15).until_not(
                    EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Loading Data')]"))
                )
                print("✅ 'Loading Data' sudah hilang.")
            except:
                print("⚠️ Timeout: Tulisan 'Loading Data' masih ada.")

            # Cek dan handle jika ada popup "Format latitude tidak valid"
            try:
                error_popup = WebDriverWait(driver, 3).until(
                    EC.visibility_of_element_located((By.CLASS_NAME, "swal2-popup"))
                )
                error_text = error_popup.find_element(By.CLASS_NAME, "swal2-html-container").text
                if "Format latitude tidak valid" in error_text:
                    print("⚠️ Popup error: Format latitude tidak valid muncul.")
                    ok_button = error_popup.find_element(By.CSS_SELECTOR, "button.swal2-confirm")
                    driver.execute_script("arguments[0].click();", ok_button)
                    print("✅ Tombol OK diklik.")
                    time.sleep(1)
            except:
                print("✅ Tidak ada popup error latitude.")


            for el in ["sumber_profiling", "catatan_profiling", "latitude", "longitude", "cek-peta"]:
                if len(driver.find_elements(By.ID, el)) > 0:
                    print(f"✅ ID ditemukan: {el}")
                else:
                    print(f"❌ ID tidak ditemukan: {el}")

            if not tunggu_loading_data_hilang(driver):
                print("🔁 Reload halaman dan ulangi proses entri...")
                retry_count += 1
                continue

            isi_input(driver, By.ID, "alamat_usaha", row.get("alamat"))
            isi_input(driver, By.ID, "sumber_profiling", row.get("Sumber"))
            isi_input(driver, By.ID, "catatan_profiling", row.get("Desk Sumber"))
            isi_input(driver, By.ID, "sls", row.get("nmsls"))

            def safe_float(val):
                try:
                    return round(float(val), 6)
                except:
                    return None

            time.sleep(1)  # Tunggu sebentar sebelum mengisi koordinat
            lat = safe_float(row.get("latitude2"))
            lon = safe_float(row.get("longitude2"))
            print(f"🌍 Akan mengisi: Latitude = {lat}, Longitude = {lon}")
            isi_input(driver, By.ID, "latitude", lat)
            isi_input(driver, By.ID, "longitude", lon)

            time.sleep(5)  # Tunggu sebentar sebelum mengisi koordinat
            try:
                safe_click(driver, By.ID, "cek-peta")
                print("🗺️ Tombol 'Cek Peta' diklik.")
            except:
                print("❌ Tombol cek-peta tidak ada")

            try:
                email = driver.find_element(By.ID, "email")
                if email.get_attribute("value").strip() == "":
                    checkbox = driver.find_element(By.ID, "check-email")
                    if checkbox.is_selected():
                        driver.execute_script("arguments[0].click();", checkbox)
                        print("📭 Uncheck email checkbox")
            except:
                print("⚠️ Email field tidak tersedia")

            safe_click(driver, By.ID, "submit-final")
            print("✅ Klik submit-final")

            try:
                WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, "confirm-consistency"))).click()
                print("✅ Konfirmasi konsistensi")
            except:
                pass

            try:
                WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.ID, "ignore-consistency"))).click()
                print("⚠️ Ignore konsistensi")
            except:
                pass

            try:
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.swal2-confirm.btn.btn-primary"))).click()
                print("🎯 Final submit berhasil")
            except:
                print("❌ Gagal submit terakhir")

            ws[f"X{excel_row}"] = "Aji"
            for cell in ws[excel_row]:
                cell.font = Font(bold=True)
            wb.save("usaha.xlsx")
            print(f"📝 Baris ke-{excel_row} selesai diproses.")
            success = True

        except Exception as e:
            print(f"❌ Error umum: {e}")
            time.sleep(5)
            
    if not success:
        print(f"❌ Gagal memproses baris {excel_row} setelah {max_retry} kali.")
    i += 1
    
print("✅ Semua baris selesai diproses.")
driver.quit()